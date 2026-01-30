import argparse
from datetime import datetime, timedelta
from html import unescape
import re
from typing import Dict, List, Optional, Set, Tuple
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill
import requests
import schedule
import time

API_BASE = "https://api.hel.fi/linkedevents/v1/"
LOCATION_SEARCH_TERM = "Savoy Teatteri"
TARGET_YEAR = 2026
WORKBOOK_PATH = "2026 Savoy-Tapahtumat Automation.xlsx"
SHEET_NAME = "Savoy Tapahtumat 2026"
REQUEST_TIMEOUT = 15
LOCAL_TZ = ZoneInfo("Europe/Helsinki")
FUZZY_THRESHOLD = 0.65
TOKEN_OVERLAP_THRESHOLD = 0.6
MAX_DURATION_HOURS = 8  # sanity cap to avoid bogus long durations
HEADERS = [
    "Event_Title",
    "Event Date and Time",
    "Flyygeli",
    "Parking",
    "Extra_Charges",
    "Väliaika",
    "Kesto",
    "Varaus_Alkaa",
    "Varaus_Loppuu",
    "Esitys_Alkaa",
    "Esitys_Loppuu",
    "weekday",
    "date",
    "ContactName",
    "e-mail",
    "Phone Number",
    "Serial Teams",
    "Event Type",
    "Event Description",
    "Event Image",
    "event_dt_iso",
    "Notified",
]
WEEKDAY_MAP = {
    0: "ma",
    1: "ti",
    2: "ke",
    3: "to",
    4: "pe",
    5: "la",
    6: "su",
}
CHANGED_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
# For "maybe-related" new events (same day but only 1 exact word overlap)
RELATED_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MANUAL_DEFAULTS = {
    "Flyygeli": "---",
    "Parking": "---",
    "Extra_Charges": "---",
}


def _normalize_title(title: str) -> str:
    normalized = unescape(title).lower()
    normalized = normalized.replace("–", "-").replace("—", "-")
    normalized = re.sub(r"[^a-z0-9äöå\-\s]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _tokenize_title(normalized_title: str) -> Set[str]:
    """Tokenize normalized title, filtering out very short tokens (< 3 chars)."""
    tokens = normalized_title.split()
    return {t for t in tokens if len(t) >= 3}


def _token_overlap_ratio(title1: str, title2: str) -> float:
    """Calculate token overlap ratio between two normalized titles."""
    tokens1 = _tokenize_title(title1)
    tokens2 = _tokenize_title(title2)
    if not tokens1 or not tokens2:
        return 0.0
    common = tokens1 & tokens2
    shorter_len = min(len(tokens1), len(tokens2))
    if shorter_len == 0:
        return 0.0
    return len(common) / shorter_len


def fetch_place_id(search_term: str = LOCATION_SEARCH_TERM) -> Optional[str]:
    response = requests.get(
        f"{API_BASE}search/",
        params={"q": search_term, "type": "place"},
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    for item in response.json().get("data", []):
        if item.get("resource_type") != "place":
            continue
        name = item.get("name") or {}
        label = name.get("fi") or name.get("sv") or name.get("en") or ""
        if "savoy" in label.lower():
            return item["id"]
    return None


def fetch_events(location_id: str, year: int) -> List[dict]:
    events: List[dict] = []
    url = f"{API_BASE}event/"
    params = {
        "location": location_id,
        "start": f"{year}-01-01T00:00:00",
        "end": f"{year}-12-31T23:59:59",
        "sort": "start_time",
        "page_size": 100,
    }

    while url:
        response = requests.get(url, params=params, timeout=REQUEST_TIMEOUT)
        if response.status_code != 200:
            raise RuntimeError(
                f"Linked Events API error {response.status_code}: {response.text}"
            )
        payload = response.json()
        events.extend(payload.get("data", []))
        url = payload.get("next")
        params = None

    filtered = [e for e in events if _event_in_year(e, year)]
    return filtered


def _event_in_year(event: dict, year: int) -> bool:
    start = event.get("start_time")
    if not start:
        return False
    try:
        return datetime.fromisoformat(start.replace("Z", "+00:00")).year == year
    except ValueError:
        return False


def _pick_translated(field: Optional[Dict[str, str]]) -> str:
    if not field:
        return ""
    return field.get("fi") or field.get("sv") or field.get("en") or ""


def _strip_html(text: str) -> str:
    clean = re.sub(r"<[^>]+>", " ", unescape(text))
    clean = re.sub(r"\s+", " ", clean, flags=re.MULTILINE).strip()
    return clean


def _parse_to_local(dt_str: Optional[str]) -> Optional[datetime]:
    if not dt_str:
        return None
    try:
        parsed = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.astimezone(LOCAL_TZ)


def _format_duration(start: datetime, end: Optional[datetime]) -> Optional[str]:
    if not end:
        return None
    delta: timedelta = end - start
    minutes = int(delta.total_seconds() // 60)
    if minutes <= 0:
        return None
    hours, minutes = divmod(minutes, 60)
    if hours > MAX_DURATION_HOURS:
        return None
    if hours == 0 and minutes == 0:
        return None
    return f"{hours} h {minutes} min"


def _parse_duration_and_intermission(text: str) -> Tuple[Optional[str], Optional[str]]:
    if not text:
        return None, None
    lower = text.lower()
    # detect intermission - check for explicit mentions first
    intermission = None
    # Explicit "no intermission" patterns
    if any(phrase in lower for phrase in ["no intermission", "without intermission", "ei väliaikaa", "ei väliaika"]):
        intermission = "ei"
    # Explicit "has intermission" patterns (Finnish)
    elif any(phrase in lower for phrase in [
        "sisältää väliajan",
        "sisältäen väliajan",
        "sisältää väliaika",
        "sisältäen väliaika",
        "väliajan sisään",
        "väliaika sisään",
        "sisältää väliajan",
    ]):
        intermission = "on"
    # Generic mentions (fallback - less specific)
    elif "intermission" in lower or "interval" in lower or "väliaika" in lower:
        intermission = "on"

    # detect duration (hours/minutes patterns)
    # Examples: "1,5 h", "1.5 h", "1 h 30 min", "90 min", "1h 10min"
    duration = None
    # Replace comma with dot for decimal parsing
    normalized = lower.replace(",", ".")
    # pattern for decimal hours
    m = re.search(r"(\d+(?:\.\d+)?)\s*h", normalized)
    if m:
        hours_float = float(m.group(1))
        total_minutes = int(round(hours_float * 60))
        if total_minutes <= 0 or total_minutes > MAX_DURATION_HOURS * 60:
            total_minutes = 0
        h, minutes = divmod(total_minutes, 60)
        duration = f"{h} h {minutes} min" if total_minutes > 0 else None
    else:
        # pattern for explicit hours + minutes
        m = re.search(r"(\d+)\s*h[^0-9]+(\d+)\s*min", normalized)
        if m:
            h = int(m.group(1))
            minutes = int(m.group(2))
            if h <= MAX_DURATION_HOURS:
                duration = f"{h} h {minutes} min"
        else:
            # pattern for minutes only
            m = re.search(r"(\d+)\s*min", normalized)
            if m:
                minutes = int(m.group(1))
                h, minutes = divmod(minutes, 60)
                if h <= MAX_DURATION_HOURS:
                    duration = f"{h} h {minutes} min"

    return duration, intermission


def _extract_event_type(event: dict) -> str:
    for keyword in event.get("keywords", []):
        label = _pick_translated(keyword.get("name"))
        if label:
            return label
    return ""


def event_to_row(event: dict) -> Optional[Dict[str, Optional[str]]]:
    start_local = _parse_to_local(event.get("start_time"))
    if not start_local:
        return None
    end_local = _parse_to_local(event.get("end_time"))

    name = _pick_translated(event.get("name"))
    description_raw = _pick_translated(event.get("description"))
    description = _strip_html(description_raw) if description_raw else ""
    images = event.get("images") or []
    image_url = images[0].get("url") if images else ""
    event_type = _extract_event_type(event)

    event_date_and_time = (
        f"{WEEKDAY_MAP[start_local.weekday()]} "
        f"{start_local.day}.{start_local.month}.{start_local.year} "
        f"klo {start_local.strftime('%H.%M')}"
    )

    row: Dict[str, Optional[str]] = {
        "Event_Title": name,
        "Event Date and Time": event_date_and_time,
        "Esitys_Alkaa": start_local.strftime("%H:%M"),
        "weekday": WEEKDAY_MAP[start_local.weekday()],
        "date": start_local.strftime("%d.%m"),
        "Event Type": event_type,
        "Event Description": description,
        "Event Image": image_url,
        "event_dt_iso": start_local.isoformat(),
    }

    if end_local:
        row["Esitys_Loppuu"] = end_local.strftime("%H:%M")

    duration = _format_duration(start_local, end_local)
    parsed_duration, intermission = _parse_duration_and_intermission(description)
    
    if duration:
        row["Kesto"] = duration
    elif parsed_duration:
        row["Kesto"] = parsed_duration

    # Always set intermission if detected in description, regardless of duration source
    if intermission:
        row["Väliaika"] = intermission

    return row


def ensure_header(ws) -> List[str]:
    """
    Ensure the worksheet has a header row containing canonical column names.
    This function is resilient to reordered or renamed columns: it will
    detect an existing header row (within the first 10 rows), normalize any
    known synonyms to canonical header names, and append any missing canonical
    headers at the end instead of overwriting the sheet layout.
    Returns the list of header names (in left-to-right order) as strings.
    """
    # helper to normalize header text for matching
    def _norm(s: object) -> str:
        if s is None:
            return ""
        return re.sub(r"[^a-z0-9]", "", str(s).lower())

    # synonyms mapping (normalized -> canonical)
    synonyms = {
        _norm("event title"): "Event_Title",
        _norm("title"): "Event_Title",
        _norm("event_title"): "Event_Title",
        _norm("eventdateandtime"): "Event Date and Time",
        _norm("event date and time"): "Event Date and Time",
        _norm("eventdatetime"): "Event Date and Time",
        _norm("event_dt_iso"): "event_dt_iso",
        _norm("eventdtiso"): "event_dt_iso",
        _norm("esitysalkaa"): "Esitys_Alkaa",
        _norm("esitysloppuu"): "Esitys_Loppuu",
        _norm("kesto"): "Kesto",
        _norm("valiaika"): "Väliaika",
        _norm("date"): "date",
        _norm("weekday"): "weekday",
        _norm("eventdescription"): "Event Description",
        _norm("event_image"): "Event Image",
        _norm("eventtype"): "Event Type",
    }

    # Search for the header row in the first 10 rows (or fewer)
    max_search_row = min(10, ws.max_row or 10)
    best_row_idx = None
    best_matches = 0
    for r in range(1, max_search_row + 1):
        row = [cell.value for cell in ws[r]]
        matches = 0
        for v in row:
            n = _norm(v)
            if n in synonyms or v in HEADERS:
                matches += 1
        if matches > best_matches:
            best_matches = matches
            best_row_idx = r

    # If no reasonable header found, create one at row 1
    if not best_row_idx or best_matches == 0:
        # preserve any existing content by inserting new row 1
        ws.insert_rows(1)
        for idx, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=idx).value = h
        return HEADERS

    # Normalize headers in the detected row
    header_row = best_row_idx
    current_headers = [cell.value for cell in ws[header_row]]
    normalized_headers: List[str] = []
    for idx, val in enumerate(current_headers, start=1):
        n = _norm(val)
        if n in synonyms:
            canon = synonyms[n]
            ws.cell(row=header_row, column=idx).value = canon
            normalized_headers.append(canon)
        elif val in HEADERS:
            normalized_headers.append(val)
        else:
            # keep unknown header as-is to preserve user layout
            normalized_headers.append(val if val is not None else "")

    # Append any missing canonical headers at the end
    missing = [h for h in HEADERS if h not in normalized_headers]
    if missing:
        start_col = len(normalized_headers) + 1
        for offset, h in enumerate(missing):
            ws.cell(row=header_row, column=start_col + offset).value = h
            normalized_headers.append(h)

    # If header row was not the first row, move it to row 1 to keep
    # the rest of the code consistent (headers are expected in ws[1]).
    if header_row != 1:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(1)}{header_row}:{openpyxl.utils.get_column_letter(len(normalized_headers))}{header_row}",
            rows=-(header_row - 1),
        )
    return normalized_headers


def build_lookup(ws, headers: List[str]) -> Tuple[Dict[str, int], Dict[Tuple[str, str], int]]:
    col_index = {name: idx for idx, name in enumerate(headers)}
    by_iso: Dict[str, int] = {}
    by_title_date: Dict[Tuple[str, str], int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row]
        iso_val = values[col_index["event_dt_iso"]] if "event_dt_iso" in col_index else None
        title = values[col_index["Event_Title"]] if "Event_Title" in col_index else None
        date_str = (
            values[col_index["Event Date and Time"]] if "Event Date and Time" in col_index else None
        )
        if iso_val:
            norm = _normalize_iso_key(iso_val)
            if norm:
                by_iso[norm] = row_idx
            else:
                by_iso[str(iso_val)] = row_idx
        if title and date_str:
            by_title_date[(str(title), str(date_str))] = row_idx
    return by_iso, by_title_date


def _parse_iso_to_dt(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _normalize_iso_key(value: object) -> Optional[str]:
    """Normalize various datetime representations into a canonical ISO string in LOCAL_TZ.
    Handles datetime objects, ISO strings with/without timezone, and common variants.
    Returns None if parsing fails.
    """
    if value is None or value == "":
        return None
    # If it's already a datetime
    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is None:
            # treat naive datetimes as local timezone
            dt = dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(LOCAL_TZ).isoformat()

    s = str(value).strip()
    # Try direct ISO parse
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(LOCAL_TZ).isoformat()
    except Exception:
        pass

    # Try common variant 'YYYY-MM-DD HH:MM:SS' by replacing space with 'T'
    try:
        dt = datetime.fromisoformat(s.replace(" ", "T"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(LOCAL_TZ).isoformat()
    except Exception:
        return None


def _normalize_iso_key(value: object) -> Optional[str]:
    """Normalize various datetime representations (Excel datetimes, naive strings,
    ISO strings with/without timezone) into a canonical ISO string in LOCAL_TZ.
    Returns None if parsing fails.
    """
    if value is None or value == "":
        return None
    # datetime object
    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(LOCAL_TZ).isoformat()

    s = str(value).strip()
    # Try direct ISO parse
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(LOCAL_TZ).isoformat()
    except Exception:
        pass

    # Try common variant 'YYYY-MM-DD HH:MM:SS' by replacing space with T
    try:
        dt = datetime.fromisoformat(s.replace(" ", "T"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(LOCAL_TZ).isoformat()
    except Exception:
        return None


def _parse_day_key_from_event_date_and_time(value: Optional[str]) -> Optional[str]:
    """
    Parse 'Event Date and Time' strings like:
      'pe 9.1.2026 klo 19.00'
    and return ISO day key 'YYYY-MM-DD'.
    """
    if not value:
        return None
    s = str(value)
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
    if not m:
        return None
    day = int(m.group(1))
    month = int(m.group(2))
    year = int(m.group(3))
    try:
        return datetime(year, month, day).date().isoformat()
    except ValueError:
        return None


def _coerce_day_key(value: object) -> Optional[str]:
    """
    Coerce a cell value into an ISO day key 'YYYY-MM-DD' when possible.
    Supports datetime/date-like objects and strings.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    # openpyxl can return date as datetime.datetime or datetime.date depending on file
    if hasattr(value, "isoformat") and not isinstance(value, str):
        try:
            # datetime.date has .isoformat() and no .date()
            return value.isoformat()
        except Exception:
            pass
    # Try ISO datetime / ISO date strings
    parsed = _parse_iso_to_dt(str(value))
    if parsed:
        return parsed.date().isoformat()
    # Try dd.mm.yyyy embedded in strings
    parsed2 = _parse_day_key_from_event_date_and_time(str(value))
    if parsed2:
        return parsed2
    return None


def build_day_titles(ws, headers: List[str]) -> Dict[str, List[Tuple[str, int]]]:
    col_index = {name: idx for idx, name in enumerate(headers)}
    date_idx = col_index.get("date")
    title_idx = col_index.get("Event_Title")
    event_dt_idx = col_index.get("event_dt_iso")
    event_date_time_idx = col_index.get("Event Date and Time")
    if date_idx is None or title_idx is None:
        return {}

    by_day: Dict[str, List[Tuple[str, int]]] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        date_val = row[date_idx].value
        title_val = row[title_idx].value
        if not title_val:
            continue

        day_key = None
        if event_dt_idx is not None:
            day_key = _coerce_day_key(row[event_dt_idx].value)
        if not day_key and event_date_time_idx is not None:
            day_key = _parse_day_key_from_event_date_and_time(row[event_date_time_idx].value)
        if not day_key:
            day_key = _coerce_day_key(date_val)

        if day_key:
            by_day.setdefault(day_key, []).append((str(title_val), row_idx))
    return by_day


def _find_insert_row(ws, iso_dt: Optional[datetime]) -> int:
    """
    Find the row index (1-based) where a new event should be inserted
    to keep ascending order by event_dt_iso. Defaults to append.
    """
    if not iso_dt:
        return ws.max_row + 1

    headers = ensure_header(ws)
    col_index = {name: idx for idx, name in enumerate(headers)}
    iso_col = col_index.get("event_dt_iso")
    if iso_col is None:
        return ws.max_row + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        existing_iso = row[iso_col].value
        existing_dt = _parse_iso_to_dt(existing_iso)
        if existing_dt and iso_dt < existing_dt:
            return row_idx
    return ws.max_row + 1


def _set_row_strike(ws, row_idx: int, strike: bool) -> None:
    for cell in ws[row_idx]:
        current_font = cell.font or Font()
        cell.font = current_font.copy(strike=strike)


def upsert_events(ws, events: List[dict]) -> Tuple[int, int, int, Set[str], Set[str]]:
    headers = ensure_header(ws)
    col_index = {name: idx for idx, name in enumerate(headers)}
    by_iso, by_title_date = build_lookup(ws, headers)
    by_day_titles = build_day_titles(ws, headers)

    added = 0
    updated = 0
    unchanged = 0
    active_iso: Set[str] = set()
    cancelled_iso: Set[str] = set()
    now = datetime.now(tz=LOCAL_TZ)

    conflicts: List[dict] = []

    for event in events:
        row_data = event_to_row(event)
        if not row_data:
            continue
        iso_key = row_data.get("event_dt_iso")
        norm_iso = _normalize_iso_key(iso_key) if iso_key else None
        if norm_iso:
            active_iso.add(norm_iso)
            if event.get("event_status") == "EventCancelled":
                cancelled_iso.add(norm_iso)

        iso_dt = _parse_iso_to_dt(norm_iso)
        # Skip touching past events entirely
        if iso_dt and iso_dt < now:
            continue

        target_row = None
        lookup_key = (
            row_data.get("Event_Title"),
            row_data.get("Event Date and Time"),
        )
        day_key = iso_dt.date().isoformat() if iso_dt else None

        related_one_word = False

        if norm_iso and norm_iso in by_iso:
            target_row = by_iso[norm_iso]
        elif lookup_key in by_title_date:
            target_row = by_title_date[lookup_key]
        best_candidate = None
        best_match_count = 0
        if day_key and day_key in by_day_titles:
            incoming_title = row_data.get("Event_Title") or ""
            incoming_norm = _normalize_title(incoming_title)
            incoming_words = set(incoming_norm.split())
            # Filter out very short words (less than 2 chars) as they're likely not meaningful
            incoming_words = {w for w in incoming_words if len(w) >= 2}

            max_matching_words = 0
            for existing_title, row_idx in by_day_titles[day_key]:
                existing_norm = _normalize_title(existing_title)
                existing_words = set(existing_norm.split())
                existing_words = {w for w in existing_words if len(w) >= 2}
                
                # Count exact word matches
                matching_words = incoming_words & existing_words
                if len(matching_words) > max_matching_words:
                    max_matching_words = len(matching_words)
                    best_candidate = (existing_title, row_idx)
                if len(matching_words) >= 2:
                    target_row = row_idx
                    break
            if not target_row and max_matching_words >= 1:
                related_one_word = True
                # record a conflict for manual review: incoming vs best_candidate
                if best_candidate:
                    candidate_title, candidate_row = best_candidate
                    cand_values = [cell.value for cell in ws[candidate_row - 1]]
                    cand_iso = None
                    try:
                        # read event_dt_iso cell value if present
                        iso_idx = col_index.get("event_dt_iso")
                        if iso_idx is not None:
                            cand_iso = ws.cell(row=candidate_row, column=iso_idx + 1).value
                    except Exception:
                        cand_iso = None

                    conflicts.append({
                        "incoming_title": incoming_title,
                        "incoming_iso": row_data.get("event_dt_iso"),
                        "incoming_display": row_data.get("Event Date and Time"),
                        "candidate_row": candidate_row,
                        "candidate_title": candidate_title,
                        "candidate_iso": cand_iso,
                        "incoming_row": row_data,
                    })

        if target_row:
            # Update the existing row's event_dt_iso cell to canonical value
            unchanged += 1
            try:
                iso_col_idx = col_index.get("event_dt_iso")
                if iso_col_idx is not None and norm_iso:
                    ws.cell(row=target_row, column=iso_col_idx + 1).value = norm_iso
                    _set_row_strike(ws, target_row, False)
            except Exception:
                pass
            if day_key and row_data.get("Event_Title"):
                by_day_titles.setdefault(str(day_key), []).append((str(row_data["Event_Title"]), target_row))
        else:
            new_row = []
            for col in headers:
                val = row_data.get(col)
                if val is None and col in MANUAL_DEFAULTS:
                    val = MANUAL_DEFAULTS[col]
                new_row.append(val)
            insert_at = _find_insert_row(ws, iso_dt)
            ws.insert_rows(insert_at)
            fill_for_new = RELATED_FILL if related_one_word else CHANGED_FILL
            for idx, value in enumerate(new_row, start=1):
                cell = ws.cell(row=insert_at, column=idx)
                cell.value = value
                if value not in (None, "") and HEADERS[idx - 1] in row_data:
                    cell.fill = fill_for_new
            _set_row_strike(ws, insert_at, False)
            if day_key and row_data.get("Event_Title"):
                by_day_titles.setdefault(str(day_key), []).append((str(row_data["Event_Title"]), insert_at))
            added += 1

    return added, updated, unchanged, active_iso, cancelled_iso, conflicts


def mark_cancellations(ws, active_iso: Set[str], cancelled_iso: Set[str]) -> int:
    headers = ensure_header(ws)
    col_index = {name: idx for idx, name in enumerate(headers)}
    iso_col = col_index.get("event_dt_iso")
    if iso_col is None:
        return 0

    now = datetime.now(tz=LOCAL_TZ)
    cancelled = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        iso_val = row[iso_col].value
        if not iso_val:
            continue
        norm = _normalize_iso_key(iso_val)
        if not norm:
            continue
        iso_dt = _parse_iso_to_dt(norm)
        if not iso_dt:
            continue

        if iso_dt < now:
            continue

        if norm in cancelled_iso:
            _set_row_strike(ws, row_idx, True)
            cancelled += 1
        elif norm not in active_iso:
            _set_row_strike(ws, row_idx, True)
            cancelled += 1
        else:
            _set_row_strike(ws, row_idx, False)

    return cancelled


def interactive_resolve(ws, conflicts: List[dict]) -> Tuple[int, int, int]:
    """Interactively resolve conflicts.

    Returns a tuple (kept, replaced, merged)
    """
    headers = ensure_header(ws)
    col_index = {name: idx for idx, name in enumerate(headers)}
    kept = 0
    replaced = 0
    merged = 0

    for c in conflicts:
        print("\nPossible duplicate found:")
        print(f"Incoming: {c.get('incoming_title')} — {c.get('incoming_display')} ({c.get('incoming_iso')})")
        print(f"Candidate row {c.get('candidate_row')}: {c.get('candidate_title')} — {c.get('candidate_iso')}")
        print("Action? [k]eep existing / [r]eplace with incoming / [m]erge (update time) / [s]kip")
        while True:
            choice = input("Choice (k/r/m/s): ").strip().lower() or "k"
            if choice in ("k", "r", "m", "s"):
                break
            print("Please enter k, r, m or s")

        candidate_row = c.get("candidate_row")
        incoming_row = c.get("incoming_row") or {}

        if choice == "k" or choice == "s":
            kept += 1
            continue

        if choice == "r":
            # replace all known header columns with incoming values
            for idx, col in enumerate(headers, start=1):
                val = incoming_row.get(col)
                if val is None and col in MANUAL_DEFAULTS:
                    val = MANUAL_DEFAULTS[col]
                ws.cell(row=candidate_row, column=idx).value = val
            _set_row_strike(ws, candidate_row, False)
            replaced += 1
            continue

        if choice == "m":
            # update time-related fields only
            iso_idx = col_index.get("event_dt_iso")
            if iso_idx is not None and incoming_row.get("event_dt_iso"):
                ws.cell(row=candidate_row, column=iso_idx + 1).value = incoming_row.get("event_dt_iso")
            start_idx = col_index.get("Esitys_Alkaa")
            end_idx = col_index.get("Esitys_Loppuu")
            if start_idx is not None and incoming_row.get("Esitys_Alkaa"):
                ws.cell(row=candidate_row, column=start_idx + 1).value = incoming_row.get("Esitys_Alkaa")
            if end_idx is not None and incoming_row.get("Esitys_Loppuu"):
                ws.cell(row=candidate_row, column=end_idx + 1).value = incoming_row.get("Esitys_Loppuu")
            _set_row_strike(ws, candidate_row, False)
            merged += 1

    return kept, replaced, merged


def sync_events(year: int = TARGET_YEAR, interactive: bool = False, auto_resolve: str = "ask") -> None:
    # default non-interactive; interactive flag handled in main()
    place_id = fetch_place_id()
    if not place_id:
        print("Savoy Teatteri place not found in Linked Events.")
        return

    events = fetch_events(place_id, year)
    if not events:
        print(f"No events returned for {year}. Workbook left unchanged.")
        return

    load_failed = False
    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH, keep_links=True)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    except PermissionError:
        # Workbook is open/locked by another process (e.g. Excel). Create a new workbook
        # in memory and mark that loading failed so we save to an alternate path later.
        wb = openpyxl.Workbook()
        load_failed = True

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.create_sheet(SHEET_NAME)

    added, updated, unchanged, active_iso, cancelled_iso, conflicts = upsert_events(ws, events)
    # Resolve or record conflicts depending on mode
    if conflicts:
        if interactive:
            kept, replaced, merged = interactive_resolve(ws, conflicts)
            print(f"Interactive resolution: kept {kept}, replaced {replaced}, merged {merged}")
        elif auto_resolve == "prefer_incoming":
            # Replace candidate rows with incoming phrasing/values
            headers = ensure_header(ws)
            col_index = {name: idx for idx, name in enumerate(headers)}
            for c in conflicts:
                candidate_row = c.get("candidate_row")
                incoming_row = c.get("incoming_row") or {}
                # replace title and time-related fields
                title_idx = col_index.get("Event_Title")
                iso_idx = col_index.get("event_dt_iso")
                evt_disp_idx = col_index.get("Event Date and Time")
                start_idx = col_index.get("Esitys_Alkaa")
                end_idx = col_index.get("Esitys_Loppuu")
                if title_idx is not None and incoming_row.get("Event_Title"):
                    ws.cell(row=candidate_row, column=title_idx + 1).value = incoming_row.get("Event_Title")
                if iso_idx is not None and incoming_row.get("event_dt_iso"):
                    ws.cell(row=candidate_row, column=iso_idx + 1).value = incoming_row.get("event_dt_iso")
                if evt_disp_idx is not None and incoming_row.get("Event Date and Time"):
                    ws.cell(row=candidate_row, column=evt_disp_idx + 1).value = incoming_row.get("Event Date and Time")
                if start_idx is not None and incoming_row.get("Esitys_Alkaa"):
                    ws.cell(row=candidate_row, column=start_idx + 1).value = incoming_row.get("Esitys_Alkaa")
                if end_idx is not None and incoming_row.get("Esitys_Loppuu"):
                    ws.cell(row=candidate_row, column=end_idx + 1).value = incoming_row.get("Esitys_Loppuu")
                _set_row_strike(ws, candidate_row, False)
            print(f"Auto-resolved {len(conflicts)} conflicts (prefer_incoming).")
        elif auto_resolve == "prefer_existing":
            print(f"Left {len(conflicts)} conflicts unchanged (prefer_existing). See 'Conflicts' sheet for details.")
        else:
            # write conflicts to sheet 'Conflicts' for review
            if "Conflicts" in wb.sheetnames:
                wb.remove(wb["Conflicts"])
            cws = wb.create_sheet("Conflicts")
            cws.append(["Incoming Title", "Incoming ISO", "Incoming Display", "Candidate Row", "Candidate Title", "Candidate ISO"])
            for c in conflicts:
                cws.append([c.get("incoming_title"), c.get("incoming_iso"), c.get("incoming_display"), c.get("candidate_row"), c.get("candidate_title"), c.get("candidate_iso")])
    cancelled = mark_cancellations(ws, active_iso, cancelled_iso)
    try:
        # If loading failed due to a locked file, avoid overwriting and save to alternate.
        if load_failed:
            raise PermissionError
        wb.save(WORKBOOK_PATH)
        save_path = WORKBOOK_PATH
    except PermissionError:
        # Workbook is likely open in Excel; fall back to writing a new file to avoid data loss.
        alt = WORKBOOK_PATH.replace(".xlsx", "") + ".updated.xlsx"
        wb.save(alt)
        save_path = alt

    print(
        f"Synced {len(events)} events for {year}: "
        f"added {added}, changed {updated}, unchanged {unchanged}, cancelled {cancelled}."
        f" Saved to: {save_path}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync Savoy events into Excel.")
    parser.add_argument("--year", type=int, default=TARGET_YEAR, help="Target year to sync.")
    parser.add_argument(
        "--watch",
        action="store_true",
        help="Keep running and sync every hour.",
    )
    parser.add_argument(
        "--interval-minutes",
        type=int,
        default=60,
        help="Polling interval in minutes when --watch is set.",
    )
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Run interactively and prompt for resolving ambiguous matches.",
    )
    parser.add_argument(
        "--auto-resolve",
        choices=["ask", "prefer_incoming", "prefer_existing"],
        default="ask",
        help="Automatically resolve ambiguous matches: 'prefer_incoming' will replace existing rows with incoming phrasing; 'prefer_existing' keeps sheet as-is; 'ask' (default) leaves for manual review or interactive mode.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    def job():
        try:
            sync_events(args.year, interactive=args.interactive, auto_resolve=args.auto_resolve)
        except Exception as exc:
            print(f"Sync failed: {exc}")

    job()
    if args.watch:
        schedule.every(args.interval_minutes).minutes.do(job)
        while True:
            schedule.run_pending()
            time.sleep(1)


if __name__ == "__main__":
    main()